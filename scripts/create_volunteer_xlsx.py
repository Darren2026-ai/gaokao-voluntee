from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = '30个志愿清单'

# 列宽
sheet.column_dimensions['A'].width = 8
sheet.column_dimensions['B'].width = 18
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 30
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 10

# 标题
title_font = Font(bold=True, size=11, color='FFFFFF')
title_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ['序号', '学校', '专业组', '具体专业', '分数', '位次']
for col, h in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col, value=h)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 冲刺数据
data = [
    (1,'上交大','101','计算机、人工智能、网络工程',687,150),
    (2,'清华','02','计算机、人工智能、自动化',698,20),
    (3,'北大','02','计算机、电子信息类、生物科学',690,50),
    (4,'浙大','101','计算机、人工智能、数据科学',675,600),
    (5,'华科','101','计算机、人工智能、数据科学',653,3600),
    (6,'华南理工','101','计算机、人工智能、软件工程',638,6800),
    (7,'西交大','101','计算机、人工智能、软件工程',648,4200),
    (8,'电子科大','101','计算机、人工智能、网络空间安全',641,5900),
    (9,'复旦','101','计算机、网络空间安全、数据科学',681,450),
    (10,'南科大','101','理学试验班、工学试验班',648,4200),
    (11,'中大','214','计算机、人工智能、数据科学',642,5500),
    (12,'武大','101','计算机、人工智能、网络工程',643,5200),
    (13,'南大','101','计算机、信息管理、物联网工程',665,1400),
    (14,'华东师大','104','计算机师范、人工智能',639,6500),
    (15,'北师大','101','数学与应用数学师范、数学类',645,4800),
    (16,'同济','101','计算机、网络工程、软件工程',660,1800),
    (17,'山大','101','计算机、人工智能、软件工程',638,6800),
    (18,'大工','101','计算机、人工智能、网络工程',636,7800),
    (19,'北师大珠海','210','数学类、理科试验班、计算机',625,11765),
    (20,'暨大','103','计算机类、电子信息类',612,17800),
    (21,'华南师范','101','数学与应用数学师范、物理学师范等',591,22000),
    (22,'南理工','101','计算机、人工智能、信息工程',618,15500),
    (23,'哈工大深圳','101','计算机、人工智能、通信工程',635,8200),
    (24,'重大','101','计算机、人工智能、软件工程',625,11500),
    (25,'天大','101','计算机、软件工程、网络工程',643,5200),
    (26,'深大','101','数学类、统计学、应用数学',606,20000),
    (27,'广州大学','101','数学与应用数学师范、数学类',590,23000),
    (28,'广工','101','数学类、统计学、应用数学',572,30000),
    (29,'广医','101','临床医学、口腔医学、基础医学',580,32000),
    (30,'山大威海','101','计算机、网络工程、数据科学',588,31500),
]

# 填充颜色
冲_fill = PatternFill(start_color='F3E5AB', end_color='F3E5AB', fill_type='solid')
稳_fill = PatternFill(start_color='C5E1A5', end_color='C5E1A5', fill_type='solid')
保_fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')

for idx, row_data in enumerate(data, 2):
    if idx <= 11:
        fill = 冲_fill
    elif idx <= 23:
        fill = 稳_fill
    else:
        fill = 保_fill
    
    for col, val in enumerate(row_data, 1):
        cell = sheet.cell(row=idx, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

wb.save('/Users/daibin/WorkBuddy/20260317214030/gaokao-volunteer/30-COLLEGES-VOLUNTEER-LIST.xlsx')
print('✅ Excel表格已生成!')
